from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 노란색 배경색 지정
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def compare_and_highlight_in_both_files(file1, file2):
    # 엑셀 파일 로드
    wb1 = load_workbook(file1)
    wb2 = load_workbook(file2)

    # 각 워크시트 비교
    for sheet_name in wb1.sheetnames:
        ws1 = wb1[sheet_name]
        ws2 = wb2[sheet_name]

        # 각 셀 비교
        for row in range(1, ws1.max_row + 1):
            for col in range(1, ws1.max_column + 1):
                cell1 = ws1.cell(row=row, column=col).value
                cell2 = ws2.cell(row=row, column=col).value
                
                # 셀 값이 다를 경우 각 파일에 노란색으로 칠하기
                if cell1 != cell2:
                    ws1.cell(row=row, column=col).fill = fill  # file1에 노란색 칠하기
                    ws2.cell(row=row, column=col).fill = fill  # file2에 노란색 칠하기
                

    # 변경된 내용을 각각의 파일로 저장
    wb1.save(file1)
    wb2.save(file2)
    print("각 파일의 차이가 있는 셀에 노란색이 적용되었습니다.")

# 사용
file1 = "F:/v3.3.1.xlsx"
file2 = "F:/유효성 에러 파일.xlsx"

compare_and_highlight_in_both_files(file1, file2)
