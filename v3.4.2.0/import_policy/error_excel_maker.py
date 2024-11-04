'''
이 코드는 [정책 Import] 시 각 정책 유효성 에러 발생, 중복 에러 발생, 정상 작동되도록 만드는 3개의 엑셀 파일을 생성하는 코드입니다.
ONE 버전에 맞게 오류 엑셀 파일을 작성하시어 사용하시기 바랍니다.
버전에 맞게 한 번 작성해두면 앞으로 계속 사용하실 수 있습니다.

테스트 코드를 작성하기 위해 참고한 부분
http://175.113.83.14/projects/rms_30_global/wiki/Excel_ImportExport_TestCase

만약, 더 추가해야 하는 부분이 있다면 댓글로 알려주시면 감사하겠습니다.

※ 에러 정리해둔 엑셀 파일 작성 시, 주의 사항
- 정책 이름, 'Pass', 'Finish', 'All Finish'는 사용 불가
('Pass'는 통과를, 'Finish'는 각 정책 해당 컬럼의 데이터 끝을, 'All Finish'는 모든 정책이 끝났음을 의미함.)
'''

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Border, Side
from sheet_list import export_sheet_names, each_sheet

# 오류 사항 저장해둔 엑셀 파일 불러오기(ONE 버전에 따라 다른 파일이 들어가야 함.)
error_file_path = 'F:/정책 Import_Export_v3.1.2.xlsx'
# RMS에서 원하는 버전 [정책 Export]한 엑셀 파일 불러오기
file_path = 'F:/detect-policy.xlsx'


error_wb = openpyxl.load_workbook(error_file_path)

# 특정 행의 마지막 열을 찾는 함수
def get_last_column_in_row(sheet, row):
    for col in range(1, sheet.max_column + 1):
        if sheet.cell(row=row, column=col).value is None:  # 값이 None인 경우
            return col - 1  # 마지막 유효한 열 반환
    return sheet.max_column  # 모든 열에 값이 있는 경우

# 모든 시트에 대해 반복
for kind_of_accack_sheet in error_wb.sheetnames:
    sheet = error_wb[kind_of_accack_sheet]  # 현재 시트 가져오기
    row = 1 # 행(1, 2, 3)
    
    # 각 정책 찾기
    while True:
        policy_name_cell_value = sheet.cell(row=row, column=1).value
        
        # 모든 정책 확인 시 반복 끝내기
        if policy_name_cell_value == 'All Finish':  # 'All Finish' 문자열을 만난 경우
            break
        
        # 정책 이름이 export_sheet_names에 포함되어 있는 경우
        if policy_name_cell_value in export_sheet_names:
            row += 1  # 다음 행으로 이동
            last_col = get_last_column_in_row(sheet, row)  # 현재 행에서 마지막 열 찾기

            # A열부터 현재 행의 마지막 열까지 반복
            for col in range(1, last_col + 1):
                each_policy_cell_value = sheet.cell(row=row, column=col).value  # 현재 정책 셀 값
                # print(each_policy_cell_value)  # 정책 셀 값 출력

                index = 2  # 인덱스 초기화 (여기서 인덱스는 항목의 순서를 나타냄)
                each_value_row = row + 1  # 다음 행을 지정 (현재 정책 항목의 바로 아래 행부터 시작)

                while True:  # 각 항목을 반복
                    each_value = sheet.cell(row=each_value_row, column=col).value  # 현재 행의 해당 열 값 읽기

                    if each_value == 'Finish':  # 'Finish' 문자열을 만난 경우 루프 종료
                        # print(1)
                        break

                    if each_value == 'None':  # 'None' 문자열을 None으로 변환
                        each_value = None
                    print(each_policy_cell_value)
                    # 각 시트의 데이터 구조에 정책 값 추가
                    each_sheet[kind_of_accack_sheet][policy_name_cell_value][each_policy_cell_value]['index'].append(index)
                    each_sheet[kind_of_accack_sheet][policy_name_cell_value][each_policy_cell_value]['change'].append(each_value)

                    print(f'policy_name : {each_policy_cell_value}, index : {index}, value : {each_value}')  # 인덱스와 값 출력
                    
                    index += 1  # 인덱스 증가
                    each_value_row += 1  # 다음 행으로 이동


                    if index > 20:  # 인덱스가 20을 초과하면 루프 종료
                        break

        
        row += 1  # 다음 행으로 이동 

        if (row > sheet.max_row):  # max_row로 현재 시트의 최대 행 수 확인
            print("최대 행을 초과했습니다. 루프를 종료합니다.")
            break


    
## RMS에서 다운로드 받은 파일을 바꾸는 코드 시작
cell_colors = ["FFFF80", "F9D28B", "BBEBBC"] # 바꾼 부분 표시할 바탕 색깔(각각 유효성 에러, 중복 에러, 정상 파일에 대한 색깔이다.)
file_names = ["유효성 에러 파일", "중복 에러 파일", "정상 파일"]

for file_index, (kind_of_attack, kind_of_attack_list) in enumerate(each_sheet.items()):
    wb = openpyxl.load_workbook(file_path) # for문 안에서 새로 열지 않으면 이전 수정했던 파일 위에 더 수정하는 상황이 된다.
    for sheet_index, (sheet_name, value) in enumerate(kind_of_attack_list.items()):
        if sheet_name in wb.sheetnames: # 해당 시트 이름이 엑셀 파일에 있을 경우
            sheet = wb[sheet_name]  # 시트 이름 또는 인덱스로 시트 선택
        else: # 없다면
            continue # 그 다음 이어서 시작
        
        print(sheet_name)

        count = 1 # 시작 기준을 변경하기 위해 만든 변수(시간복잡도 문제로 이렇게 함, 이렇게 되면 엑셀 정책 행의 순서와 딕셔너리의 순서가 같아야 정상 작동함.)
        for column_name in each_sheet[kind_of_attack][sheet_name]: # 여기서 each_sheet[kind_of_attack][sheet_name]은 각 정책의 딕셔너리를 의미한다.
            # 첫 번째 행(헤더)에서 '공격명'이 포함된 열 번호 찾기
            for col in sheet.iter_cols(count, sheet.max_column, 1, 1):  # 첫 번째 행만 검사
                header = col[0].value
                
                if header and column_name in str(header):  # 헤더가 '공격명'을 포함하는지 확인
                    count += 1 # 원하는 요소를 찾아서 1 증가
                    column_letter = get_column_letter(col[0].column)
                    
                    for i in range(len(each_sheet[kind_of_attack][sheet_name][column_name]['index'])):
                        if each_sheet[kind_of_attack][sheet_name][column_name]['change'][i] == 'Pass':
                            continue
                        
                        # 셀 값 변경
                        sheet[f'{column_letter}{each_sheet[kind_of_attack][sheet_name][column_name]['index'][i]}'] = each_sheet[kind_of_attack][sheet_name][column_name]['change'][i]
                        
                        # 배경색 노란색으로 변경
                        yellow_fill = PatternFill(start_color=cell_colors[file_index], end_color=cell_colors[file_index], fill_type="solid")
                        
                        # 굵은(Bold) 글씨체
                        bold_font = Font(bold=True)
                        
                        # 테두리 설정 (두께는 thin, 색상은 검정)
                        thin_border = Border(
                            left=Side(border_style="thin", color="000000"),
                            right=Side(border_style="thin", color="000000"),
                            top=Side(border_style="thin", color="000000"),
                            bottom=Side(border_style="thin", color="000000")
                        )

                        # 셀 스타일 적용
                        cell = sheet[f'{column_letter}{each_sheet[kind_of_attack][sheet_name][column_name]['index'][i]}']
                        cell.font = bold_font         # 글자 굵게
                        cell.fill = yellow_fill       # 셀 색깔
                        cell.border = thin_border     # 테두리
                        
                    break
            else: # break에 걸리지 않으면 실행됨.(즉, count열부터 1행 끝열(데이터가 있는 가장 마지막 열)까지 해당 요소가 없으면)
                print(f"\n'{column_name}'이 포함된 열을 찾을 수 없습니다.")
        

    # 다른 이름으로 저장하기
    new_file_path = f'F:/{file_names[file_index]}.xlsx'  # 새 파일 이름(원하는 경로로 지정)
    wb.save(new_file_path)  # 수정된 내용을 다른 이름으로 저장
    
    # 다시 열었다가 저장(이렇게 안 하면 RMS에서 오류가 나는 경우 꽤 있음.)
    wb = openpyxl.load_workbook(new_file_path)
    wb.save(new_file_path)
