'''
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 엑셀 파일 열기
wb = load_workbook("F:/detect-policy.xlsx")  # 파일 경로에 맞게 수정
ws = wb.active  # 첫 번째 시트를 선택합니다. 다른 시트를 선택하려면 ws = wb["시트이름"]

# 공백 셀 배경 색상 지정
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 노란색 배경

# 모든 시트를 순회하면서 공백인 셀 찾기
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]  # 각 시트를 선택
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                cell.fill = fill  # 공백인 셀의 배경 색상을 변경

# 변경 사항 저장
wb.save("원래 RMS.xlsx")  # 변경 사항을 새 파일로 저장
wb.close()
'''
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 엑셀 파일 열기
wb = load_workbook("F:/유효성 에러 파일.xlsx")  # 파일 경로에 맞게 수정
ws = wb.active  # 첫 번째 시트를 선택합니다. 다른 시트를 선택하려면 ws = wb["시트이름"]

# 공백 셀 배경 색상 지정
fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 노란색 배경

# 모든 시트를 순회하면서 공백인 셀 찾기
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]  # 각 시트를 선택
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                cell.fill = fill  # 공백인 셀의 배경 색상을 변경

# 변경 사항 저장
wb.save("코드로 만든 파일.xlsx")  # 변경 사항을 새 파일로 저장
wb.close()